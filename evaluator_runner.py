# -*- coding: utf-8 -*-
"""
evaluator_runner.py
-------------------
Utility script for H&R Hub to (1) generate chatbot answers from Evaluator questions
and (2) build an evaluation workbook with fidelity metrics.

Requirements (already used in your project):
- openpyxl
- scikit-learn
- Python 3.9+ recommended

How to use
==========
1) Place this file in the SAME folder as your main project file: "H&R Hub.py".
   (The file name has a space and an ampersand; that's OK. We will import it via importlib.)

2) Ensure this folder contains the subfolder "EVALUATOR" with:
   - "Evaluator questions.xlsx"  (first sheet is used)

3) Ensure the environment variable OPENAI_API_KEY is set, because we reuse
   the existing `call_openai_generate` function defined in "H&R Hub.py".
   Example (Linux/macOS):
       export OPENAI_API_KEY="sk-..."
   Example (Windows PowerShell):
       setx OPENAI_API_KEY "sk-..."

4) Run from a terminal (inside the folder that contains "H&R Hub.py"):

    python evaluator_runner.py           # Runs BOTH steps (answers + metrics)
    # or, for individual steps:
    python evaluator_runner.py answers   # Only generates EVALUATOR/Chatbot Answers.xlsx
    python evaluator_runner.py metrics   # Only generates EVALUATOR/Metrics H&R Hub Chatbot.xlsx
    python evaluator_runner.py metrics_llm   # NEW: Generates EVALUATOR/Metrics H&R Hub Chatbot (LLM).xlsx

What this script does
=====================
Part 1 (Answers):
- Loads the first sheet of EVALUATOR/Evaluator questions.xlsx.
- Keeps all existing formatting from the template.
- Ensures there is a column exactly to the RIGHT of "Questions" named "Chatbot Answer".
  If a column named "Ideal Answer" exists, it is REPLACED by "Chatbot Answer" and MOVED
  next to "Questions" if necessary. Other columns remain in place.
- Sends each non-empty question (row by row, preserving order) to the SAME chat pipeline used
  by your app: reuse `load_corpus`, `build_index`, `rank_chunks`, and `call_openai_generate`
  from "H&R Hub.py" (no reconfiguration). We do NOT rewrite anything.
- If an answer attempt fails, it retries once. If it fails again, writes
  "Error al generar respuesta" in that row.
- Answers are written in English (this is already enforced by your `call_openai_generate` function).
- Sets text wrapping ON for the "Chatbot Answer" column and adjusts row heights for readability.
- Saves as EVALUATOR/Chatbot Answers.xlsx.

Part 2 (Metrics):
- Loads the template EVALUATOR/Evaluator questions.xlsx (first sheet) and PRESERVES all columns.
- Keeps the original "Ideal Answer" column (renamed in the output header to
  "Ideal Answer (Ground Truth)" for clarity) and ADDS a "Chatbot Answer" column (do NOT replace).
  Matching is performed by "Questions" text, and for duplicates, by row order.
- Computes:
    * Fidelity %  (semantic-like score using a blend of word- and char-level TF‑IDF cosine)
    * High Fidelity (TRUE if Fidelity % > 85, else FALSE)
    * Keyword Overlap % (Jaccard overlap of top TF‑IDF keywords)
    * Length Difference % (approximate difference in token length)
  If the chatbot answer is missing, leaves the cell blank and sets Fidelity % to 0.
- Adds a short "Notes" column for brief observations.
- Wraps text for "Ideal Answer (Ground Truth)", "Chatbot Answer", and "Notes";
  adjusts row heights; and keeps other columns unchanged.
- Saves as EVALUATOR/Metrics H&R Hub Chatbot.xlsx.
- Prints a summary to console and also writes a small "Summary" sheet in the metrics workbook.

Important
=========
- We import your main module from the literal file "H&R Hub.py" (with space and &). Do not rename it.
- We call ONLY the existing functions (no reconfiguration): load_corpus, build_index,
  rank_chunks, call_openai_generate.

New (LLM-based Metrics)
=======================
- Command: `python evaluator_runner.py metrics_llm`
- Uses an OpenAI model as evaluator to compare each Ideal Answer vs Chatbot Answer and
  outputs: "Sufficiently Correct" (TRUE/FALSE), "Score (1–10)", and a short "Evaluator Notes".
- Saves as EVALUATOR/"Metrics H&R Hub Chatbot (LLM).xlsx".
"""

import os
import sys
import math
import copy
import logging
import json
import datetime
import importlib.util
from dataclasses import dataclass
from typing import Dict, List, Optional, Tuple

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from sklearn.feature_extraction.text import TfidfVectorizer, ENGLISH_STOP_WORDS
from sklearn.metrics.pairwise import cosine_similarity
from openai import OpenAI


# ---------------------------- Logging ---------------------------------
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)s | %(message)s",
    datefmt="%H:%M:%S",
)
logger = logging.getLogger("evaluator_runner")


# ---------------------------- Utilities ---------------------------------
# Common header synonyms
QUESTION_HEADERS = [
    "Questions",
    "Question",
    "Pregunta",
    "Preguntas",
    "Pregunta(s)",
]
IDEAL_HEADERS = [
    "Ideal Answer",
    "Ideal Answer (Ground Truth)",
    "Ground Truth",
    "Respuesta ideal",
    "Respuesta Ideal",
]
CHATBOT_HEADERS = [
    "Chatbot Answer",
    "Respuesta Chatbot",
    "Chatbot",
]
def normalize_ws(text: Optional[str]) -> str:
    if text is None:
        return ""
    return " ".join(str(text).split())


def safe_lower(s: Optional[str]) -> str:
    return normalize_ws(s).lower()


def find_header_col(ws, wanted_names: List[str]) -> Optional[int]:
    """
    Find a column index (1-based) whose header value (row 1) matches any of the wanted names (case-insensitive).
    """
    headers: Dict[str, int] = {}
    header_list: List[Tuple[int, str]] = []
    for col in range(1, ws.max_column + 1):
        raw = ws.cell(row=1, column=col).value
        hdr = safe_lower(raw)
        headers[hdr] = col
        header_list.append((col, hdr))
    # exact match first
    for name in wanted_names:
        nm = name.strip().lower()
        if nm in headers:
            return headers[nm]
    # relaxed contains match, but ignore empty headers
    for col, hdr in header_list:
        if not hdr:
            continue
        for name in wanted_names:
            nm = name.strip().lower()
            if not nm:
                continue
            if nm in hdr or hdr in nm:
                return col
    return None


def get_header_value(ws, col: int) -> str:
    try:
        return normalize_ws(ws.cell(row=1, column=col).value)
    except Exception:
        return ""


def select_first_sheet_with_header(wb, wanted_names: List[str]):
    for ws in wb.worksheets:
        col = find_header_col(ws, wanted_names)
        if col is not None:
            return ws
    return wb.worksheets[0] if wb.worksheets else None


def copy_cell_style(src_cell, dst_cell):
    dst_cell.font = copy.copy(src_cell.font)
    dst_cell.fill = copy.copy(src_cell.fill)
    dst_cell.border = copy.copy(src_cell.border)
    # Preserve horizontal alignment if set on source; ensure wrap for target when needed elsewhere
    dst_cell.alignment = copy.copy(src_cell.alignment)
    dst_cell.number_format = src_cell.number_format
    dst_cell.protection = copy.copy(src_cell.protection)


def set_col_wrap(ws, col_idx: int, vertical: str = "top"):
    for r in range(1, ws.max_row + 1):
        c = ws.cell(row=r, column=col_idx)
        # Keep horizontal alignment if present; enforce wrap + vertical top
        current = c.alignment or Alignment()
        c.alignment = Alignment(
            wrap_text=True,
            horizontal=current.horizontal,
            vertical=vertical
        )


def estimate_row_height(text: str, approx_chars_per_line: float) -> float:
    """
    Estimate a reasonable row height given text length and approximate characters per line.
    """
    if not text:
        return 15.0  # default-ish
    lines = text.count("\n") + max(1, math.ceil(len(text) / max(1.0, approx_chars_per_line)))
    # baseline 15 points; add ~12 points per extra wrapped line; clamp
    height = 15.0 + (lines - 1) * 12.0
    return min(height, 220.0)


def column_width(ws, col_idx: int) -> float:
    letter = get_column_letter(col_idx)
    cd = ws.column_dimensions.get(letter)
    return float(cd.width) if (cd and cd.width) else 10.0  # Excel default ≈ 8.43; use 10 as safer default


def set_column_width(ws, col_idx: int, width: float):
    letter = get_column_letter(col_idx)
    ws.column_dimensions[letter].width = width


def insert_blank_column(ws, insert_at: int, copy_width_from: Optional[int] = None) -> None:
    ws.insert_cols(insert_at, amount=1)
    if copy_width_from:
        set_column_width(ws, insert_at, column_width(ws, copy_width_from))


def move_or_create_chatbot_column(ws, q_col: int) -> int:
    """
    Ensures a column named "Chatbot Answer" exists *exactly to the right* of Questions.
    If "Ideal Answer" exists, it will be renamed to "Chatbot Answer" and MOVED next to Questions if needed.
    Other columns are preserved.
    Returns the final 1-based index of "Chatbot Answer".
    """
    dest_col = q_col + 1
    ia_col = find_header_col(ws, IDEAL_HEADERS)
    ch_col = find_header_col(ws, CHATBOT_HEADERS)

    # If Chatbot Answer already exists
    if ch_col is not None:
        if ch_col != dest_col:
            # Move it next to Questions: insert at dest and copy, then delete original
            logger.info("Moving existing 'Chatbot Answer' column from %s to %s", ch_col, dest_col)
            insert_blank_column(ws, dest_col, copy_width_from=ch_col if ch_col < dest_col else None)
            # Adjust source index if it was shifted by insertion
            src = ch_col + 1 if ch_col >= dest_col else ch_col
            for r in range(1, ws.max_row + 1):
                src_cell = ws.cell(row=r, column=src)
                dst_cell = ws.cell(row=r, column=dest_col)
                dst_cell.value = src_cell.value
                copy_cell_style(src_cell, dst_cell)
            # Delete old position
            ws.delete_cols(src, 1)
            ch_col = dest_col
        # Ensure header text is exactly "Chatbot Answer"
        ws.cell(row=1, column=ch_col).value = "Chatbot Answer"
        return ch_col

    # If Ideal Answer exists, turn it into Chatbot Answer at dest_col
    if ia_col is not None:
        logger.info("Transforming 'Ideal Answer' (col %s) into 'Chatbot Answer' at col %s", ia_col, dest_col)
        insert_blank_column(ws, dest_col, copy_width_from=ia_col if ia_col < dest_col else None)
        # After insertion, if ia_col >= dest_col, it has shifted by +1
        src = ia_col + 1 if ia_col >= dest_col else ia_col
        # Copy cell values + styles
        for r in range(1, ws.max_row + 1):
            src_cell = ws.cell(row=r, column=src)
            dst_cell = ws.cell(row=r, column=dest_col)
            dst_cell.value = src_cell.value if r == 1 else None  # We'll write answers later; keep header later
            copy_cell_style(src_cell, dst_cell)
        # Set exact header name
        ws.cell(row=1, column=dest_col).value = "Chatbot Answer"
        # Delete old 'Ideal Answer' column
        ws.delete_cols(src, 1)
        return dest_col

    # Neither Chatbot Answer nor Ideal Answer exists: create a new column to the right of Questions
    logger.info("Inserting new 'Chatbot Answer' column at col %s", dest_col)
    insert_blank_column(ws, dest_col, copy_width_from=q_col)
    ws.cell(row=1, column=dest_col).value = "Chatbot Answer"
    # Set a comfortable width if Questions is narrow
    current_width = column_width(ws, dest_col)
    if current_width < 40.0:
        set_column_width(ws, dest_col, 70.0)
    return dest_col


# ---------------------------- Import H&R Hub main module ----------------
def import_hr_hub_module(project_root: Path):
    module_path = project_root / "H&R Hub.py"
    if not module_path.exists():
        raise FileNotFoundError(f"Cannot find main file at: {module_path}")
    spec = importlib.util.spec_from_file_location("hr_hub_main", str(module_path))
    hr_hub = importlib.util.module_from_spec(spec)  # type: ignore
    assert spec and spec.loader, "Failed to prepare module spec for H&R Hub.py"
    spec.loader.exec_module(hr_hub)  # type: ignore
    # Sanity check required functions
    required = ["load_corpus", "build_index", "rank_chunks", "call_openai_generate"]
    for fn in required:
        if not hasattr(hr_hub, fn):
            raise AttributeError(f"'H&R Hub.py' is missing required function: {fn}")
    return hr_hub


# ---------------------------- Part 1: Generate Chatbot Answers ----------
def generate_chatbot_answers(project_root: Path, evaluator_dir: Path) -> Path:
    """
    Reads EVALUATOR/Evaluator questions.xlsx and writes EVALUATOR/Chatbot Answers.xlsx.
    """
    hr_hub = import_hr_hub_module(project_root)

    # Build RAG context (reuse your exact pipeline)
    logger.info("Loading corpus and building index using H&R Hub functions...")
    chunks = hr_hub.load_corpus(str(project_root))
    vectorizer, matrix = hr_hub.build_index(chunks)

    # Load template workbook
    src_path = evaluator_dir / "Evaluator questions.xlsx"
    if not src_path.exists():
        raise FileNotFoundError(f"Missing input workbook: {src_path}")

    wb = load_workbook(filename=str(src_path))
    ws = select_first_sheet_with_header(wb, QUESTION_HEADERS)
    if ws is None:
        raise ValueError("No worksheet found in the workbook.")

    q_col = find_header_col(ws, QUESTION_HEADERS)
    if q_col is None:
        raise ValueError("The workbook must contain a 'Questions' column in row 1.")

    logger.info("Detected 'Questions' column at index %s (header='%s')", q_col, get_header_value(ws, q_col))

    # Ensure the destination column exists and is correctly positioned
    chat_col = move_or_create_chatbot_column(ws, q_col=q_col)
    # Ensure wrap text for "Chatbot Answer"
    set_col_wrap(ws, chat_col)

    # Approx characters per line for row height estimation (based on column width)
    approx_cpl = max(18.0, column_width(ws, chat_col) * 1.1)

    # Iterate rows (row 1 is header)
    max_row = ws.max_row
    total = 0
    failed = 0

    logger.info("Generating answers row by row...")
    for r in range(2, max_row + 1):
        q_val = ws.cell(row=r, column=q_col).value
        question = normalize_ws(q_val)
        if not question:
            # Ignore empty rows
            continue

        total += 1
        answer_text: Optional[str] = None

        try:
            ranked = hr_hub.rank_chunks(question, vectorizer, matrix, chunks, top_k=25)
            answer_text = hr_hub.call_openai_generate(question, ranked, max_sentences=5)
            if not answer_text or not normalize_ws(answer_text):
                # Retry once
                ranked = hr_hub.rank_chunks(question, vectorizer, matrix, chunks, top_k=25)
                answer_text = hr_hub.call_openai_generate(question, ranked, max_sentences=5)
        except Exception as e:
            logger.warning("Error generating answer (row %s): %s", r, e)

        if not answer_text or not normalize_ws(answer_text):
            answer_text = "Error al generar respuesta"

        # Write answer
        cell = ws.cell(row=r, column=chat_col)
        cell.value = answer_text

        # Wrap text (already set for column) and adjust row height
        row_h = estimate_row_height(str(answer_text), approx_chars_per_line=approx_cpl)
        ws.row_dimensions[r].height = row_h

    # Save as new workbook with timestamp
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = evaluator_dir / f"Chatbot Answers_{timestamp}.xlsx"
    wb.save(str(out_path))
    logger.info("Saved: %s", out_path)
    logger.info("Processed questions: %s | Failures: %s", total, failed)
    return out_path


# ---------------------------- Part 2: Build Metrics Workbook ------------
@dataclass
class Pairing:
    row_eval: int
    row_chat: Optional[int]
    question: str
    ideal: str
    chatbot: str


def tokenize(text: str) -> List[str]:
    return [t for t in "".join([c.lower() if c.isalnum() or c.isspace() else " " for c in text]).split() if t]


def top_keywords(texts: List[str], top_k: int = 12) -> List[List[str]]:
    """
    Extract top keywords per text using TF-IDF (1-2 grams) ignoring sklearn's English stopwords.
    Returns a list of keyword lists aligned with input texts.
    """
    try:
        vec = TfidfVectorizer(ngram_range=(1, 2), stop_words="english")
        X = vec.fit_transform(texts)
        feats = vec.get_feature_names_out()
        out: List[List[str]] = []
        for i in range(X.shape[0]):
            row = X.getrow(i)
            if row.nnz == 0:
                out.append([])
                continue
            arr = row.toarray()[0]
            idxs = arr.argsort()[::-1]
            kws: List[str] = []
            for j in idxs:
                tok = feats[j]
                if len(tok) < 3:
                    continue
                if tok in ENGLISH_STOP_WORDS:
                    continue
                kws.append(tok)
                if len(kws) >= top_k:
                    break
            out.append(kws)
        return out
    except Exception:
        return [[] for _ in texts]


def jaccard(a: List[str], b: List[str]) -> float:
    A, B = set(a), set(b)
    if not A and not B:
        return 0.0
    return len(A & B) / len(A | B)


def cosine_pair_word_and_char(a: str, b: str) -> Tuple[float, float, float]:
    """
    Returns (cos_word, cos_char, blended) where blended = 0.6*cos_word + 0.4*cos_char.
    """
    a2, b2 = normalize_ws(a), normalize_ws(b)
    try:
        v_word = TfidfVectorizer(ngram_range=(1, 2), stop_words="english")
        M = v_word.fit_transform([a2, b2])
        cw = float(cosine_similarity(M[0], M[1])[0, 0])
    except Exception:
        cw = 0.0
    try:
        v_char = TfidfVectorizer(analyzer="char_wb", ngram_range=(3, 5))
        M2 = v_char.fit_transform([a2, b2])
        cc = float(cosine_similarity(M2[0], M2[1])[0, 0])
    except Exception:
        cc = 0.0
    blended = 0.6 * cw + 0.4 * cc
    return cw, cc, blended


def compute_metrics(ideal: str, chatbot: str) -> Tuple[float, float, float, str]:
    """
    Returns: Fidelity %, Keyword Overlap %, Length Difference %, notes
    """
    ideal = normalize_ws(ideal)
    chatbot = normalize_ws(chatbot)

    if chatbot == "":
        return 0.0, 0.0, 100.0, "Missing chatbot answer"

    # Similarity (semantic-like) via blended cosine
    cw, cc, blended = cosine_pair_word_and_char(ideal, chatbot)

    # Length difference (token-based), in percent
    toks_i = tokenize(ideal)
    toks_c = tokenize(chatbot)
    if len(toks_i) == 0:
        length_diff_pct = 100.0 if len(toks_c) > 0 else 0.0
    else:
        length_diff_pct = 100.0 * abs(len(toks_i) - len(toks_c)) / max(1, len(toks_i))

    # Keywords overlap
    kw_i, kw_c = top_keywords([ideal, chatbot], top_k=12)
    ko = 100.0 * jaccard(kw_i, kw_c)

    # Combine into Fidelity % with a mild length penalty
    # (kept simple and transparent; values clamped to [0, 100])
    penalty = max(0.5, 1.0 - 0.5 * (length_diff_pct / 100.0))  # strong length diffs reduce score down to 50%
    fidelity = max(0.0, min(100.0, 100.0 * blended * penalty))

    # Notes
    notes = ""
    if fidelity < 30.0:
        notes = "Low similarity"
    elif length_diff_pct > 60.0 and fidelity < 85.0:
        notes = "Large length gap"
    elif ko < 15.0 and fidelity < 85.0:
        notes = "Low keyword overlap"

    return round(fidelity, 1), round(ko, 1), round(length_diff_pct, 1), notes


def load_rows(ws, q_col: int, ia_col: Optional[int]) -> List[Tuple[int, str, str]]:
    out = []
    for r in range(2, ws.max_row + 1):
        q = normalize_ws(ws.cell(row=r, column=q_col).value)
        if not q:
            continue
        ideal = normalize_ws(ws.cell(row=r, column=ia_col).value) if ia_col else ""
        out.append((r, q, ideal))
    return out


def load_chatbot_rows(ws, q_col: int, ch_col: int) -> List[Tuple[int, str, str]]:
    out = []
    for r in range(2, ws.max_row + 1):
        q = normalize_ws(ws.cell(row=r, column=q_col).value)
        if not q:
            continue
        ch = normalize_ws(ws.cell(row=r, column=ch_col).value)
        out.append((r, q, ch))
    return out


def match_by_question_with_order(eval_rows, chat_rows) -> List[Pairing]:
    """
    Matches rows primarily by Questions text; for duplicate questions,
    pairs by appearance order (first with first, etc.).
    Returns a list of Pairing aligned to the evaluator's order.
    """
    # Create lookup: question -> list of (row, chat)
    from collections import defaultdict, deque
    buckets = defaultdict(deque)
    for r, q, ch in chat_rows:
        buckets[q].append((r, ch))

    pairings: List[Pairing] = []
    for r_eval, q, ideal in eval_rows:
        if buckets[q]:
            r_chat, ch = buckets[q].popleft()
        else:
            r_chat, ch = None, ""
        pairings.append(Pairing(row_eval=r_eval, row_chat=r_chat, question=q, ideal=ideal, chatbot=ch))
    return pairings


def build_metrics_workbook(project_root: Path, evaluator_dir: Path) -> Path:
    # Load template workbook
    src_path = evaluator_dir / "Evaluator questions.xlsx"
    answers_path = evaluator_dir / "Chatbot Answers.xlsx"
    if not src_path.exists():
        raise FileNotFoundError(f"Missing template workbook: {src_path}")
    if not answers_path.exists():
        raise FileNotFoundError(f"Missing answers workbook: {answers_path}")

    wb = load_workbook(filename=str(src_path))
    ws = select_first_sheet_with_header(wb, QUESTION_HEADERS)  # choose sheet that has Questions
    if ws is None:
        raise ValueError("No worksheet found in the template workbook.")

    q_col = find_header_col(ws, QUESTION_HEADERS)
    if q_col is None:
        raise ValueError("Template workbook must contain a 'Questions' column in row 1.")
    ia_col = find_header_col(ws, IDEAL_HEADERS)

    logger.info("Template: 'Questions' at %s (header='%s'); 'Ideal' at %s", q_col, get_header_value(ws, q_col), ia_col)

    # Load answers workbook (we rely on Questions + Chatbot Answer)
    wb_ans = load_workbook(filename=str(answers_path))
    ws_ans = select_first_sheet_with_header(wb_ans, QUESTION_HEADERS)
    if ws_ans is None:
        raise ValueError("No worksheet found in the answers workbook.")
    q_col_ans = find_header_col(ws_ans, QUESTION_HEADERS)
    ch_col_ans = find_header_col(ws_ans, CHATBOT_HEADERS)
    if q_col_ans is None or ch_col_ans is None:
        raise ValueError("Answers workbook must contain 'Questions' and 'Chatbot Answer' headers.")

    logger.info("Answers: 'Questions' at %s (header='%s'); 'Chatbot' at %s (header='%s')", q_col_ans, get_header_value(ws_ans, q_col_ans), ch_col_ans, get_header_value(ws_ans, ch_col_ans))

    # Extract ordered rows
    eval_rows = load_rows(ws, q_col, ia_col)
    chat_rows = load_chatbot_rows(ws_ans, q_col_ans, ch_col_ans)

    # Pair rows
    pairings = match_by_question_with_order(eval_rows, chat_rows)

    # In the METRICS workbook, we keep all original columns.
    # We add "Chatbot Answer" (do NOT replace), then the metrics to the right.
    # Preferred placement: immediately to the right of the "Ideal Answer" column if it exists,
    # otherwise to the right of "Questions".
    if ia_col is not None:
        dest_chat_col = ia_col + 1
    else:
        dest_chat_col = q_col + 1

    # Insert Chatbot Answer column (empty first, we'll fill next)
    insert_blank_column(ws, dest_chat_col, copy_width_from=ia_col or q_col)
    ws.cell(row=1, column=dest_chat_col).value = "Chatbot Answer"
    set_col_wrap(ws, dest_chat_col)
    if column_width(ws, dest_chat_col) < 40.0:
        set_column_width(ws, dest_chat_col, 70.0)

    # Ensure Ideal Answer header is exactly "Ideal Answer (Ground Truth)" if IA exists
    if ia_col is not None:
        ws.cell(row=1, column=ia_col).value = "Ideal Answer (Ground Truth)"
        set_col_wrap(ws, ia_col)

    # Now append metric columns after Chatbot Answer, in this order
    labels = ["Fidelity %", "High Fidelity", "Keyword Overlap %", "Length Difference %", "Notes"]
    base = dest_chat_col + 1
    for i, lab in enumerate(labels):
        insert_blank_column(ws, base + i, copy_width_from=dest_chat_col)
        ws.cell(row=1, column=base + i).value = lab
        if lab in ("Notes",):
            set_column_width(ws, base + i, max(30.0, column_width(ws, dest_chat_col) * 0.5))
        set_col_wrap(ws, base + i)

    # Fill rows in original order
    total = 0
    high = 0
    approx_cpl_chat = max(18.0, column_width(ws, dest_chat_col) * 1.1)
    approx_cpl_ideal = max(18.0, column_width(ws, ia_col) * 1.1) if ia_col else 50.0
    for p in pairings:
        r = p.row_eval
        total += 1

        # Write Chatbot Answer
        ws.cell(row=r, column=dest_chat_col).value = p.chatbot

        # Compute metrics
        fidelity, kw_overlap, length_diff, notes = compute_metrics(p.ideal, p.chatbot)
        if fidelity > 85.0:
            high += 1

        ws.cell(row=r, column=base + 0).value = fidelity
        ws.cell(row=r, column=base + 1).value = True if fidelity > 85.0 else False
        ws.cell(row=r, column=base + 2).value = kw_overlap
        ws.cell(row=r, column=base + 3).value = length_diff
        ws.cell(row=r, column=base + 4).value = notes

        # Adjust row height using the "taller" of ideal/chatbot
        h1 = estimate_row_height(p.chatbot, approx_cpl_chat)
        h2 = estimate_row_height(p.ideal, approx_cpl_ideal)
        ws.row_dimensions[r].height = max(h1, h2)

    # Add a tiny "Summary" sheet
    if "Summary" in wb.sheetnames:
        summary_ws = wb["Summary"]
    else:
        summary_ws = wb.create_sheet("Summary")
    summary_ws["A1"] = "Total Questions"
    summary_ws["B1"] = total
    summary_ws["A2"] = "High Fidelity (> 85%)"
    summary_ws["B2"] = high
    summary_ws["A3"] = "High Fidelity %"
    summary_ws["B3"] = round(100.0 * (high / total if total else 0.0), 1)

    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = evaluator_dir / f"Metrics H&R Hub Chatbot_{timestamp}.xlsx"
    wb.save(str(out_path))
    logger.info("Saved: %s", out_path)
    logger.info("Summary — Total: %s | >85%%: %s (%.1f%%)", total, high, 100.0 * (high / total if total else 0.0))
    return out_path


# ---------------------------- Part 2 (LLM): Metrics via OpenAI ---------
def call_openai_evaluator(question: str, ideal_answer: str, chatbot_answer: str, model: str = "gpt-4o-mini") -> Tuple[Optional[bool], Optional[int], str]:
    """
    Calls an OpenAI model to judge whether the chatbot answer is sufficiently correct
    compared to the ideal answer and assigns a score from 1 to 10.
    Returns: (sufficiently_correct, score, notes). On failure, (None, None, reason).
    """
    api_key = os.environ.get("OPENAI_API_KEY", "").strip()
    if not api_key:
        return None, None, "Missing OPENAI_API_KEY"

    system_msg = (
        "You are an expert evaluator for enterprise knowledge-grounded Q&A. "
        "Assess whether the chatbot answer is sufficiently correct compared to the IDEAL answer for the same question. "
        "Be strict but fair and consider factuality, coverage, clarity, and alignment with the ideal answer. "
        "Return ONLY strict JSON with keys: sufficiently_correct (boolean), score (integer 1-10), notes (short reason)."
    )
    user_msg = (
        "Evaluate the following answers to the same question.\n\n"
        f"Question:\n{question}\n\n"
        f"IDEAL ANSWER:\n{ideal_answer}\n\n"
        f"CHATBOT ANSWER:\n{chatbot_answer}\n\n"
        "Output format (JSON only): {\n"
        "  \"sufficiently_correct(no necesita ser perfecta, basta que tenga el mismo sentido o significado)\": true|false,\n"
        "  \"score\": <integer 1-10>,\n"
        "  \"notes\": \"short reason\"\n"
        "}"
    )

    try:
        client = OpenAI(api_key=api_key)
        resp = client.chat.completions.create(
            model=model,
            messages=[
                {"role": "system", "content": system_msg},
                {"role": "user", "content": user_msg},
            ],
            temperature=0.0,
            max_tokens=300,
        )
        text = (resp.choices[0].message.content or "").strip()
    except Exception as e:
        return None, None, f"OpenAI error: {e}"

    payload = text
    if payload.startswith("```"):
        try:
            payload = payload.strip().strip("`")
            if payload.lower().startswith("json"):
                payload = payload[4:].strip()
        except Exception:
            pass
    try:
        data = json.loads(payload)
        sc = data.get("sufficiently_correct")
        score = data.get("score")
        notes = data.get("notes", "")
        if isinstance(sc, bool) and isinstance(score, int):
            score = max(1, min(10, score))
            return sc, score, str(notes)
        if isinstance(sc, bool) and isinstance(score, str) and score.isdigit():
            score_int = int(score)
            score_int = max(1, min(10, score_int))
            return sc, score_int, str(notes)
        return None, None, f"Unparseable fields in response: {text[:200]}"
    except Exception:
        return None, None, f"Invalid JSON from model: {text[:200]}"


def build_metrics_workbook_llm(project_root: Path, evaluator_dir: Path) -> Path:
    """
    Builds a metrics workbook using an OpenAI evaluator. Keeps all original columns
    from the template, adds a "Chatbot Answer" column (do NOT replace Ideal Answer), then
    adds: "Sufficiently Correct", "Score (1–10)", and "Evaluator Notes" columns.
    Saves as EVALUATOR/"Metrics H&R Hub Chatbot (LLM).xlsx".
    """
    # Load template workbook
    src_path = evaluator_dir / "Evaluator questions.xlsx"
    answers_path = evaluator_dir / "Chatbot Answers.xlsx"
    if not src_path.exists():
        raise FileNotFoundError(f"Missing template workbook: {src_path}")
    if not answers_path.exists():
        raise FileNotFoundError(f"Missing answers workbook: {answers_path}")

    wb = load_workbook(filename=str(src_path))
    ws = wb.worksheets[0]

    q_col = find_header_col(ws, ["Questions"])
    if q_col is None:
        raise ValueError("Template workbook must contain a 'Questions' header in row 1.")
    ia_col = find_header_col(ws, ["Ideal Answer", "Ideal Answer (Ground Truth)", "Ground Truth"])

    # Load answers workbook
    wb_ans = load_workbook(filename=str(answers_path))
    ws_ans = wb_ans.worksheets[0]
    q_col_ans = find_header_col(ws_ans, ["Questions"])
    ch_col_ans = find_header_col(ws_ans, ["Chatbot Answer"])
    if q_col_ans is None or ch_col_ans is None:
        raise ValueError("Answers workbook must contain 'Questions' and 'Chatbot Answer' headers.")

    # Extract rows and pair by question with order
    eval_rows = load_rows(ws, q_col, ia_col)
    chat_rows = load_chatbot_rows(ws_ans, q_col_ans, ch_col_ans)
    pairings = match_by_question_with_order(eval_rows, chat_rows)

    # Place Chatbot Answer next to Ideal Answer (if present) or next to Questions
    if ia_col is not None:
        dest_chat_col = ia_col + 1
    else:
        dest_chat_col = q_col + 1

    insert_blank_column(ws, dest_chat_col, copy_width_from=ia_col or q_col)
    ws.cell(row=1, column=dest_chat_col).value = "Chatbot Answer"
    set_col_wrap(ws, dest_chat_col)
    if column_width(ws, dest_chat_col) < 40.0:
        set_column_width(ws, dest_chat_col, 70.0)

    if ia_col is not None:
        ws.cell(row=1, column=ia_col).value = "Ideal Answer (Ground Truth)"
        set_col_wrap(ws, ia_col)

    # Append LLM metric columns
    labels = ["Sufficiently Correct", "Score (1–10)", "Evaluator Notes"]
    base = dest_chat_col + 1
    for i, lab in enumerate(labels):
        insert_blank_column(ws, base + i, copy_width_from=dest_chat_col)
        ws.cell(row=1, column=base + i).value = lab
        if lab in ("Evaluator Notes",):
            set_column_width(ws, base + i, max(30.0, column_width(ws, dest_chat_col) * 0.5))
        set_col_wrap(ws, base + i)

    # Fill rows
    total = 0
    sufficient = 0
    sum_scores = 0
    count_scores = 0
    approx_cpl_chat = max(18.0, column_width(ws, dest_chat_col) * 1.1)
    approx_cpl_ideal = max(18.0, column_width(ws, ia_col) * 1.1) if ia_col else 50.0

    for p in pairings:
        r = p.row_eval
        total += 1

        # Write Chatbot Answer
        ws.cell(row=r, column=dest_chat_col).value = p.chatbot

        if not normalize_ws(p.chatbot):
            ws.cell(row=r, column=base + 0).value = False
            ws.cell(row=r, column=base + 1).value = None
            ws.cell(row=r, column=base + 2).value = "Missing chatbot answer"
            h1 = estimate_row_height(p.chatbot, approx_cpl_chat)
            h2 = estimate_row_height(p.ideal, approx_cpl_ideal)
            ws.row_dimensions[r].height = max(h1, h2)
            continue

        sc, score, notes = call_openai_evaluator(p.question, p.ideal, p.chatbot)

        if sc is True:
            sufficient += 1
        if score is not None:
            sum_scores += score
            count_scores += 1

        ws.cell(row=r, column=base + 0).value = sc if sc is not None else None
        ws.cell(row=r, column=base + 1).value = score if score is not None else None
        ws.cell(row=r, column=base + 2).value = notes

        h1 = estimate_row_height(p.chatbot, approx_cpl_chat)
        h2 = estimate_row_height(p.ideal, approx_cpl_ideal)
        ws.row_dimensions[r].height = max(h1, h2)

    # Summary sheet
    if "Summary" in wb.sheetnames:
        summary_ws = wb["Summary"]
        start_row = 6
    else:
        summary_ws = wb.create_sheet("Summary")
        start_row = 1
    summary_ws[f"A{start_row}"] = "Total Questions"
    summary_ws[f"B{start_row}"] = total
    summary_ws[f"A{start_row+1}"] = "Sufficiently Correct"
    summary_ws[f"B{start_row+1}"] = sufficient
    summary_ws[f"A{start_row+2}"] = "Sufficiently Correct %"
    summary_ws[f"B{start_row+2}"] = round(100.0 * (sufficient / total if total else 0.0), 1)
    summary_ws[f"A{start_row+3}"] = "Average Score (1–10)"
    avg_score = (sum_scores / count_scores) if count_scores else 0.0
    summary_ws[f"B{start_row+3}"] = round(avg_score, 2)

    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = evaluator_dir / f"Metrics H&R Hub Chatbot (LLM)_{timestamp}.xlsx"
    wb.save(str(out_path))
    logger.info("Saved: %s", out_path)
    logger.info(
        "LLM Summary — Total: %s | Sufficient: %s (%.1f%%) | Avg Score: %.2f",
        total,
        sufficient,
        100.0 * (sufficient / total if total else 0.0),
        avg_score,
    )
    return out_path
# ---------------------------- Main entry point --------------------------
def main(argv: List[str]) -> None:
    project_root = Path(__file__).resolve().parent
    evaluator_dir = project_root / "EVALUATOR"
    evaluator_dir.mkdir(parents=True, exist_ok=True)

    if len(argv) <= 1:
        # Default: run both steps
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        logger.info("Running BOTH steps: answers + metrics (Run ID: %s)", timestamp)
        ans_path = generate_chatbot_answers(project_root, evaluator_dir)  # Will use its own timestamp
        met_path = build_metrics_workbook(project_root, evaluator_dir)  # Will use its own timestamp
        logger.info("Done.\nAnswers: %s\nMetrics: %s", ans_path, met_path)
        return

    cmd = argv[1].strip().lower()
    if cmd == "answers":
        ans_path = generate_chatbot_answers(project_root, evaluator_dir)
        logger.info("Done. Answers: %s", ans_path)
    elif cmd == "metrics":
        met_path = build_metrics_workbook(project_root, evaluator_dir)
        logger.info("Done. Metrics: %s", met_path)
    elif cmd == "metrics_llm":
        met_path = build_metrics_workbook_llm(project_root, evaluator_dir)
        logger.info("Done. Metrics (LLM): %s", met_path)
    else:
        logger.error("Unknown command: %s\nUse: python evaluator_runner.py [answers|metrics]", cmd)


if __name__ == "__main__":
    main(sys.argv)
